from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q, Avg
from django.utils import timezone
from django.http import JsonResponse
from datetime import date, datetime, timedelta
from activities.models import DailyActivity
from employees.models import Employee, Company
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.urls import reverse
# import datetime
import pytz

User = get_user_model()


def is_admin_or_hr(user):
    """Check if user is admin or HR"""
    return user.is_staff or user.is_superuser


@login_required
def index_view(request):
    """Main dashboard view - works with or without employee profile"""
    today = date.today()
    
    # Get today's activity for the user
    today_activity = DailyActivity.objects.filter(
        user=request.user,
        date=today
    ).first()
    
    # Get recent activities
    recent_activities = DailyActivity.objects.filter(
        user=request.user
    ).order_by('-date')[:7]
    
    # Calculate stats
    this_week_start = today - timedelta(days=today.weekday())
    this_week_activities = DailyActivity.objects.filter(
        user=request.user,
        date__gte=this_week_start
    )
    
    stats = {
        'total_activities': recent_activities.count(),
        'completed_this_week': this_week_activities.filter(status='completed').count(),
        'on_time_this_week': this_week_activities.filter(attendance_status='on_time').count(),
        'late_this_week': this_week_activities.filter(attendance_status='late').count(),
    }
    
    # Check if user has employee profile
    has_employee_profile = hasattr(request.user, 'employee_profile')
    
    context = {
        'user': request.user,
        'today_activity': today_activity,
        'recent_activities': recent_activities,
        'stats': stats,
        'is_admin': is_admin_or_hr(request.user),
        'has_employee_profile': has_employee_profile,
    }
    
    if has_employee_profile:
        context['employee'] = request.user.employee_profile
        return render(request, 'dashboard/employee_dashboard.html', context)
    else:
        context['message'] = 'Welcome! Your employee profile is being set up by HR. You can still clock in/out.'
        return render(request, 'dashboard/simple_dashboard.html', context)


@login_required
@user_passes_test(is_admin_or_hr)
def admin_dashboard_view(request):
    """Admin dashboard with analytics"""
    
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    # Get filter parameters
    company_filter = request.GET.get('company', '')
    employee_filter = request.GET.get('employee', '')
    date_filter = request.GET.get('date_range', f'{start_of_week}_{end_of_week}')

    start_date_range = date_filter.split('_')[0] if date_filter else ""
    end_date_range = date_filter.split('_')[1] if date_filter else ""
    
    # Base queryset
    activities_qs = DailyActivity.objects.all()
    employees_qs = Employee.objects.filter(employment_status='active')
    
    # Apply company filter
    if company_filter:
        activities_qs = activities_qs.filter(user__employee_profile__company_id=company_filter)
        employees_qs = employees_qs.filter(company_id=company_filter)
    
    # Apply employee filter
    if employee_filter:
        activities_qs = activities_qs.filter(user__employee_profile__id=employee_filter)
    
    # Apply date filter
    if date_filter:
        activities_qs = activities_qs.filter(date__gte=start_date_range, date__lte=end_date_range)
    
    # Calculate statistics
    total_employees = employees_qs.count()
    
    # Calculate total active users based on filters
    active_users_filters = Q(is_active=True)

    if company_filter:
        active_users_filters &= Q(employee_profile__company_id=company_filter)
    
    if employee_filter:
        active_users_filters &= Q(employee_profile__id=employee_filter)

    total_active_users = User.objects.filter(active_users_filters).count()
    
    today_activities_filters = Q(date=today)

    if company_filter:
        today_activities_filters &= Q(user__employee_profile__id=company_filter)
    
    if employee_filter:
        today_activities_filters &= Q(user__employee_profile__company_id=employee_filter)

    # Today's attendance
    today_activities = DailyActivity.objects.filter(today_activities_filters)
    today_stats = {
        'total_expected': total_active_users,
        'checked_in': today_activities.filter(checkin_time__isnull=False).count(),
        'checked_out': today_activities.filter(checkout_time__isnull=False).count(),
        'on_time': today_activities.filter(attendance_status='on_time').count(),
        'late': today_activities.filter(attendance_status='late').count(),
        'absent': total_active_users - today_activities.filter(checkin_time__isnull=False).count(),
    }
    
    # Attendance trends
    attendance_stats = activities_qs.aggregate(
        total_activities=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        on_time=Count('id', filter=Q(attendance_status='on_time')),
        late=Count('id', filter=Q(attendance_status='late')),
        absent=Count('id', filter=Q(attendance_status='absent')),
    )
    
    # Recent activities
    recent_activities = activities_qs.order_by('-date', '-checkin_time')[:20]
    
    # Late arrivals today
    late_today = today_activities.filter(attendance_status='late').select_related('user')
    
    # Absent users today
    checked_in_user_ids = today_activities.filter(
        checkin_time__isnull=False
    ).values_list('user_id', flat=True)
    
    absent_today = User.objects.filter(is_active=True).exclude(id__in=checked_in_user_ids)
    
    if company_filter:
        absent_today = absent_today.filter(
            employee_profile__company_id=company_filter
        )
    
    if employee_filter:
        absent_today = absent_today.filter(
            employee_profile__id=employee_filter
        )
    
    # Companies for filter
    companies = Company.objects.filter(is_active=True)

    activities_data = []
    for activity in activities_qs:
        try:
            checkin_lat_str, checkin_lng_str = activity.checkin_location.split(",")
            checkout_lat_str, checkout_lng_str = activity.checkout_location.split(",")
            activities_data.append({
                "id":activity.id,
                "employee": activity.user.get_full_name() or activity.user.username,
                "lat": float(checkin_lat_str.strip()),
                "lng": float(checkin_lng_str.strip()),
                "checkout_lng": float(checkout_lng_str.strip()),
                "checkout_lat": float(checkout_lat_str.strip()),
                "time": timezone.localtime(activity.checkin_time).strftime("%Y-%m-%d %H:%M"),
                "checkout_time": timezone.localtime(activity.checkout_time).strftime("%Y-%m-%d %H:%M"),
                "status": activity.status,
                "attendance_status": activity.attendance_status,
                "detail_url": f"/employees/activities/{activity.id}/"
            })
        except ValueError:
            continue
    
    # Prepare context
    context = {
        'activities' :activities_data,
        'today_stats': today_stats,
        'attendance_stats': attendance_stats,
        'recent_activities': recent_activities,
        'late_today': late_today,
        'absent_today': absent_today,
        'companies': companies,
        'employees': employees_qs,  # For employee filter dropdown
        'selected_company': company_filter,
        'selected_employee': employee_filter,  # To keep selected employee in filter
        'selected_date_range': date_filter, # To keep selected date range in filter
        'total_employees': total_employees,
        'total_active_users': total_active_users,
    }
    
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
@user_passes_test(is_admin_or_hr)
def export_admin_dashboard(request):
    """Export filtered activity details to Excel"""
    
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    # Get filter parameters
    company_filter = request.GET.get('company', '')
    employee_filter = request.GET.get('employee', '')
    date_filter = request.GET.get('date_range', f'{start_of_week}_{end_of_week}')

    start_date_range = date_filter.split('_')[0] if date_filter else ""
    end_date_range = date_filter.split('_')[1] if date_filter else ""
    
    # Timezone for GMT+7
    tz = pytz.timezone('Asia/Jakarta')  # GMT+7
    
    # Base queryset
    activities_qs = DailyActivity.objects.all()
    
    # Apply company filter
    if company_filter:
        activities_qs = activities_qs.filter(user__employee_profile__company_id=company_filter)
    
    # Apply employee filter
    if employee_filter:
        activities_qs = activities_qs.filter(user__employee_profile__id=employee_filter)
    
    # Apply date filter    
    if date_filter:
        activities_qs = activities_qs.filter(date__gte=start_date_range, date__lte=end_date_range)
    
    # Prefetch related data for efficiency
    activities_qs = activities_qs.prefetch_related(
        'planned_activities',
        'daily_goals',
        'additional_activities',
        'user__employee_profile'
    ).order_by('-date', '-checkin_time')
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create Activity Details sheet
    wb.remove(wb.active)
    ws_activities = wb.create_sheet("Activity Details")
    
    # Define headers
    headers = [
        "Employee", "Position", "Company", "Date",
        "Check In (GMT+7)", "Check Out (GMT+7)", 
        "Status", "Attendance Status", "Work Duration",
        "Morning Problems", "Afternoon Problems", "Notes",
        "Planned Activities", "Activity Status", "Activity Priority", "Activity Reasons",
        "Daily Goals", "Goal Status", "Goal Priority", "Target", "Achieved", "Completion %", "Goal Reasons",
        "Additional Activities", "Add. Category", "Add. Status", "Add. Duration", "Add. Impact"
    ]
    ws_activities.append(headers)
    
    # Iterate through each activity and write to sheet
    for activity in activities_qs:
        # Handle check-in time
        checkin_str = ""
        if activity.checkin_time:
            if isinstance(activity.checkin_time, datetime):
                localized_in = activity.checkin_time.astimezone(tz)
                checkin_str = localized_in.strftime('%Y-%m-%d %H:%M:%S')
            else:
                naive_in = datetime.combine(activity.date, activity.checkin_time)
                localized_in = tz.localize(naive_in)
                checkin_str = localized_in.strftime('%Y-%m-%d %H:%M:%S')
        else:
            checkin_str = "N/A"
        
        # Handle check-out time
        checkout_str = ""
        if activity.checkout_time:
            if isinstance(activity.checkout_time, datetime):
                localized_out = activity.checkout_time.astimezone(tz)
                checkout_str = localized_out.strftime('%Y-%m-%d %H:%M:%S')
            else:
                naive_out = datetime.combine(activity.date, activity.checkout_time)
                localized_out = tz.localize(naive_out)
                checkout_str = localized_out.strftime('%Y-%m-%d %H:%M:%S')
        else:
            checkout_str = "N/A"
        
        # Calculate work duration
        work_duration = "N/A"
        if activity.work_duration:
            total_seconds = activity.work_duration.total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            work_duration = f"{hours}h {minutes}m"
        
        # Get employee profile
        employee = activity.user.employee_profile
        
        # Prepare lists for related activities
        planned_activities = list(activity.planned_activities.all())
        daily_goals = list(activity.daily_goals.all())
        additional_activities = list(activity.additional_activities.all())
        
        # Determine max rows needed for this activity
        max_rows = max(1, len(planned_activities), len(daily_goals), len(additional_activities))
        
        for i in range(max_rows):
            pa = planned_activities[i] if i < len(planned_activities) else None
            dg = daily_goals[i] if i < len(daily_goals) else None
            aa = additional_activities[i] if i < len(additional_activities) else None
            
            # Only include main activity details in first row
            if i == 0:
                row = [
                    activity.user.get_full_name() or activity.user.username,
                    employee.position if employee else "N/A",
                    employee.company.name if (employee and employee.company) else "N/A",
                    activity.date.strftime("%Y-%m-%d"),
                    checkin_str,
                    checkout_str,
                    activity.get_status_display(),
                    activity.get_attendance_status_display(),
                    work_duration,
                    activity.morning_problems or "",
                    activity.afternoon_problems or "",
                    activity.notes or ""
                ]
            else:
                # For additional rows, leave main activity details empty
                row = ["", "", "", "", "", "", "", "", "", "", "", ""]
            
            # Add planned activity details
            row.extend([
                pa.title if pa else "",
                pa.get_status_display() if pa else "",
                pa.get_priority_display() if pa else "",
                pa.reasons if pa else ""
            ])
            
            # Add daily goal details
            row.extend([
                dg.title if dg else "",
                dg.get_status_display() if dg else "",
                dg.get_priority_display() if dg else "",
                dg.target_value if dg else "",
                dg.achieved_value if dg else "",
                f"{dg.completion_percentage}%" if dg and dg.completion_percentage is not None else "",
                dg.reasons if dg else ""
            ])
            
            # Add additional activity details
            row.extend([
                aa.title if aa else "",
                aa.get_category_display() if aa else "",
                aa.get_status_display() if aa else "",
                str(aa.duration) if aa and aa.duration else "",
                aa.impact_on_planned_work if aa else ""
            ])
            
            ws_activities.append(row)
    
    # Apply styling to the sheet
    header_fill = PatternFill(start_color="2D6099", end_color="2D6099", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling to the first row
    for cell in ws_activities[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Apply borders to all cells
    for row in ws_activities.iter_rows(min_row=1, max_row=ws_activities.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Auto adjust column widths
    for col in ws_activities.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                value = str(cell.value) if cell.value else ""
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws_activities.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"activity_details_export_{datetime.now(tz).strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response